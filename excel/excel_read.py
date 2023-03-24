#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@Name: excel_read.py
@Auth: Sun dong
@Date: 2022/7/1-10:39
@Desc: 
@Ver : 0.0.0
"""
import random
import openpyxl
# excel = openpyxl.load_workbook('../data/login_test_account.xlsx')
# excel = openpyxl.load_workbook('../data/新大屏测试账及密码(1) 的副本.xlsx')
excel = openpyxl.load_workbook('../data/一般性群体报告.xlsx')
sheet = excel['学生名单']
# 读取excel内容，实现文件驱动自动化执行
def read_excel():
    tulpe_list = []
    # 逐行循环读取Excel数据
    for value in sheet.values:
        # 判断当前行第一列的值，是否是数字编号
        if type(value[0]) is int:
            # 将元祖装载进list
            account=value[1]
            tulpe_list.append(account)
    return tulpe_list


